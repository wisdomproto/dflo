"""Convert 성장치료 환자 리스트 xlsx to JSON for the Node importer.

Collapses per-chart metadata (first non-empty row wins) and collects every
unique visit-date with a flag for 초진 rows. The Node side consumes this
JSON so we don't need a Node-side xlsx parser.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "성장치료 환자 리스트 260421.xlsx"
OUT = HERE / "_patient_roster.json"


def non_empty(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main() -> None:
    wb = load_workbook(XLSX, read_only=True)
    ws = wb.active

    patients: dict[str, dict] = {}
    visits: dict[str, dict[str, bool]] = defaultdict(dict)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if i == 1:
            continue  # "◀ 과 목 ▶"
        chart = non_empty(row[3])
        if not chart:
            continue
        name = non_empty(row[4])
        presc_date = row[12]
        rrn = non_empty(row[13])
        phone = non_empty(row[15])
        email = non_empty(row[18])
        zipcode = non_empty(row[19])
        address = non_empty(row[20])
        init_type = non_empty(row[24])

        p = patients.setdefault(chart, {"chart_number": chart})
        if name and not p.get("name"):
            p["name"] = name
        if rrn and not p.get("rrn"):
            p["rrn"] = rrn
        if address and not p.get("address"):
            p["address"] = address
        if zipcode and not p.get("zipcode"):
            p["zipcode"] = zipcode
        if phone and not p.get("phone"):
            p["phone"] = phone
        if email and not p.get("email"):
            p["email"] = email

        # Visit date normalization
        if presc_date is None:
            continue
        if hasattr(presc_date, "date"):  # datetime
            visit_date = presc_date.date().isoformat()
        else:
            m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(presc_date))
            if not m:
                continue
            visit_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        is_first = bool(init_type and "초진" in init_type)
        prev = visits[chart].get(visit_date, False)
        visits[chart][visit_date] = prev or is_first

    payload = {
        "patients": [
            {**patients[c], "visits": [{"date": d, "is_first": f} for d, f in sorted(dates.items())]}
            for c, dates in visits.items()
            for _ in [None]  # single-iter shim
            if (patients.get(c) or True)
        ],
    }
    # The comprehension above iterates once per chart; rebuild cleanly:
    payload = {
        "patients": [
            {
                **patients[c],
                "visits": [
                    {"date": d, "is_first": f} for d, f in sorted(visits[c].items())
                ],
            }
            for c in sorted(patients)
        ]
    }

    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    total_visits = sum(len(dates) for dates in visits.values())
    print(f"Wrote {len(payload['patients'])} patients / {total_visits} visit-dates → {OUT}")


if __name__ == "__main__":
    main()
