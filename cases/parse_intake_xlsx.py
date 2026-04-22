"""Parse the Google Forms intake xlsx into structured JSON for DB import.

Input: cases/「187성장클리닉」 성장문진표 응답리스트.xlsx
Output: cases/intake_xlsx_parsed.json

Only emits raw answers — field-level coercion (예→true, 로마숫자→int, etc.)
is done by the Node importer so rules stay in one place.
"""
from __future__ import annotations

import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).parent
SRC = HERE / "「187성장클리닉」 성장문진표 응답리스트.xlsx"
OUT = HERE / "intake_xlsx_parsed.json"

# Column index (0-based) → logical field name
COLS = {
    0: "timestamp",
    1: "phone",
    2: "name",
    3: "birth_date",
    4: "birth_week_raw",
    5: "birth_weight_raw",
    6: "birth_notes",
    7: "grade",
    8: "class_height_rank",
    # 9-17: growth history 8-16 (col by age)
    18: "desired_height_raw",
    19: "father_height_raw",
    20: "mother_height_raw",
    21: "parents_interested_raw",  # 예/아니오
    22: "child_interested_raw",
    23: "past_clinic_consult_raw",
    24: "chronic_conditions",
    25: "sports_athlete_raw",
    26: "sports_event",
    27: "growth_pattern_raw",
    28: "short_stature_raw",
    29: "gender_raw",              # 남아/여아
    30: "tanner_female_raw",
    31: "tanner_male_raw",
}
GROWTH_AGES = {9: 8, 10: 9, 11: 10, 12: 11, 13: 12, 14: 13, 15: 14, 16: 15, 17: 16}


def norm(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = unicodedata.normalize("NFC", str(v)).strip()
    return s or None


def main():
    wb = openpyxl.load_workbook(str(SRC), data_only=True)
    ws = wb["응답(URL)"]
    records: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col_idx, key in COLS.items():
            row[key] = norm(ws.cell(r, col_idx + 1).value)
        growth = {}
        for col_idx, age in GROWTH_AGES.items():
            v = norm(ws.cell(r, col_idx + 1).value)
            if v:
                growth[str(age)] = v
        if growth:
            row["growth_history_raw"] = growth
        row["row_number"] = r
        # Skip fully blank rows
        if not row.get("name") or not row.get("birth_date"):
            continue
        records.append(row)
    OUT.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} records to {OUT}")


if __name__ == "__main__":
    main()
