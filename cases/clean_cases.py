import pandas as pd, json, numpy as np, re, sys, io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ─── Load growth chart ───
with open('v4/src/features/guide/data/growthChartData.json', 'r', encoding='utf-8') as f:
    gc = json.load(f)

chart = {}
for gender in ['male', 'female']:
    chart[gender] = {}
    for p in ['5th', '50th', '95th']:
        for pt in gc[gender][p]:
            m = pt['month']
            if m not in chart[gender]:
                chart[gender][m] = {}
            chart[gender][m][p] = pt['height']

def get_percentile(gender, month, height):
    g = 'male' if gender == 'M' else 'female'
    m = min(max(int(round(month)), 0), 216)
    if m not in chart[g]:
        return None
    d = chart[g][m]
    h5, h50, h95 = d['5th'], d['50th'], d['95th']
    if height <= h5:
        return max(1, 5 * height / h5)
    elif height <= h50:
        return 5 + 45 * (height - h5) / (h50 - h5)
    elif height <= h95:
        return 50 + 45 * (height - h50) / (h95 - h50)
    else:
        return min(99, 95 + 4 * (height - h95) / max(h95 - h50, 1))

def get_height_at_percentile(gender, month, pct):
    g = 'male' if gender == 'M' else 'female'
    m = min(max(int(round(month)), 0), 216)
    if m not in chart[g]:
        return None
    d = chart[g][m]
    h5, h50, h95 = d['5th'], d['50th'], d['95th']
    if pct <= 5:
        return h5 * pct / 5
    elif pct <= 50:
        return h5 + (h50 - h5) * (pct - 5) / 45
    elif pct <= 95:
        return h50 + (h95 - h50) * (pct - 50) / 45
    else:
        return h95 + (h95 - h50) * (pct - 95) / 4  # inverse of forward: pct = 95 + 4*(h-h95)/(h95-h50)

def predict_adult_height(gender, bone_age_months, current_height):
    pct = get_percentile(gender, bone_age_months, current_height)
    if pct is None:
        return None
    adult_h = get_height_at_percentile(gender, 216, pct)
    return round(adult_h, 1) if adult_h else None

def parse_bone_age_months(ba_str):
    """Parse bone age string to months. Conservative: use upper bound (older bone age = lower predicted height)."""
    if pd.isna(ba_str) or str(ba_str).strip() == '':
        return None
    s = str(ba_str).strip()
    # "13세4~5개월" → 13*12+5 (conservative: take upper bound)
    m = re.match(r'(\d+)\s*세\s*(\d+(?:\.\d+)?)\s*~\s*(\d+)\s*개월', s)
    if m:
        return int(m.group(1)) * 12 + int(float(m.group(3)))
    # "12세0개월", "13세 3개월" (no range)
    m1 = re.match(r'(\d+)\s*세\s*(\d+(?:\.\d+)?)\s*개월', s)
    if m1:
        return int(m1.group(1)) * 12 + int(float(m1.group(2)))
    # "13세 근방", "13세" → +6 (conservative: assume mid-year = older)
    m2 = re.match(r'(\d+)\s*세\s*(?:근방|정도|경)?', s)
    if m2:
        return int(m2.group(1)) * 12 + 6
    return None

def normalize_bone_age_str(ba_str):
    """Normalize bone age display: '11세3~4개월' → '11세4개월' (conservative: upper bound)"""
    if pd.isna(ba_str) or str(ba_str).strip() == '':
        return ''
    s = str(ba_str).strip()
    # "13세4~5개월" → "13세5개월"
    m = re.match(r'(\d+)\s*세\s*\d+\s*~\s*(\d+)\s*개월', s)
    if m:
        return f'{m.group(1)}세{m.group(2)}개월'
    # "13세 근방" → "13세6개월"
    m2 = re.match(r'(\d+)\s*세\s*(근방|정도|경)\s*$', s)
    if m2:
        return f'{m2.group(1)}세6개월'
    # Already clean, return as-is
    return s

def calc_age_months(dob, measure_date):
    if pd.isna(dob) or pd.isna(measure_date):
        return None
    d1 = pd.Timestamp(dob)
    d2 = pd.Timestamp(measure_date)
    total = (d2.year - d1.year) * 12 + (d2.month - d1.month)
    if d2.day < d1.day:
        total -= 1
    return total

def months_to_str(m):
    if m is None:
        return ''
    years = m // 12
    months = m % 12
    return f'{years}세{months}개월'

def parse_date_col(val):
    if pd.isna(val):
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.Timestamp(val)
    s = str(val).strip()
    if s.startswith('(') or ':' in s:  # skip notes like "(11:40AM)"
        return None
    # Handle "2025-06-04(?)"
    s = re.sub(r'\(\?\)', '', s).strip()
    try:
        return pd.Timestamp(s)
    except:
        return None

def parse_dob(val):
    if pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        return pd.Timestamp(datetime(1899, 12, 30) + timedelta(days=int(val)))
    s = str(val).strip().replace('\xa0', '')
    try:
        return pd.Timestamp(s)
    except:
        return None

def parse_num(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    s = re.sub(r'\(\?\)', '', s).strip()
    try:
        return float(s)
    except:
        return None

# ─── Read Excel ───
df = pd.read_excel('cases/치료 후기 케이스 정리 260409.xlsx', header=0)

# ─── Identify patients ───
patients = []
current = None
for i, row in df.iterrows():
    if pd.notna(row.get('차트번호')) and str(row.get('차트번호')).strip() != '':
        if current:
            patients.append(current)
        current = {
            'start_row': i,
            'category': str(row.get('카테고리', '')) if pd.notna(row.get('카테고리')) else '',
            'parent_id': str(row.get('부모 ID', '')) if pd.notna(row.get('부모 ID')) else '',
            'chart_no': str(row.get('차트번호', '')).strip(),
            'name': str(row.get('성함', '')),
            'dob': parse_dob(row.get('생년월일')),
            'gender': str(row.get('성별', '')).strip(),
            'dad_h': str(row.get('아빠 키', '')) if pd.notna(row.get('아빠 키')) else '',
            'mom_h': str(row.get('엄마 키', '')) if pd.notna(row.get('엄마 키')) else '',
            'target': str(row.get('목표키 (THT)', '')) if pd.notna(row.get('목표키 (THT)')) else '',
            'consult_memo': str(row.get('상담 메모', '')) if pd.notna(row.get('상담 메모')) else '',
            'review': str(row.get('치료 후기', '')) if pd.notna(row.get('치료 후기')) else '',
            'youtube': str(row.get('유투브 링크', '')) if pd.notna(row.get('유투브 링크')) else '',
            'rows': [i]
        }
    elif current:
        current['rows'].append(i)
if current:
    patients.append(current)

# ─── Process each patient ───
results = []
changes_log = []

for p in patients:
    measurements = []
    for ri in p['rows']:
        row = df.iloc[ri]
        date = parse_date_col(row.get('날짜'))
        height = parse_num(row.get('키(cm)'))
        weight = parse_num(row.get('몸무게(kg)'))
        ba_str_raw = str(row.get('뼈나이', '')) if pd.notna(row.get('뼈나이')) else ''
        ba_str = normalize_bone_age_str(ba_str_raw)
        ba_months = parse_bone_age_months(ba_str_raw)
        original_pah = str(row.get('뼈예상키 (THT/PAH)', '')) if pd.notna(row.get('뼈예상키 (THT/PAH)')) else ''
        memo1 = str(row.get('치료 메모1', '')) if pd.notna(row.get('치료 메모1')) else ''
        memo2 = str(row.get('치료 메모2', '')) if pd.notna(row.get('치료 메모2')) else ''

        if date is None:
            continue

        measurements.append({
            'date': date,
            'height': height,
            'weight': weight,
            'ba_str': ba_str,
            'ba_months': ba_months,
            'original_pah': original_pah,
            'memo1': memo1,
            'memo2': memo2,
        })

    # Sort by date
    measurements.sort(key=lambda x: x['date'])

    # ─── Fix actual age ───
    for m in measurements:
        m['age_months'] = calc_age_months(p['dob'], m['date'])
        m['age_str'] = months_to_str(m['age_months'])

    # ─── Interpolate missing heights ───
    for i in range(len(measurements)):
        if measurements[i]['height'] is None:
            prev_h, prev_d = None, None
            next_h, next_d = None, None
            for j in range(i-1, -1, -1):
                if measurements[j]['height'] is not None:
                    prev_h = measurements[j]['height']
                    prev_d = measurements[j]['date']
                    break
            for j in range(i+1, len(measurements)):
                if measurements[j]['height'] is not None:
                    next_h = measurements[j]['height']
                    next_d = measurements[j]['date']
                    break
            if prev_h and next_h and prev_d and next_d:
                total_days = (next_d - prev_d).days
                curr_days = (measurements[i]['date'] - prev_d).days
                ratio = curr_days / total_days if total_days > 0 else 0.5
                interp = round(prev_h + (next_h - prev_h) * ratio, 1)
                changes_log.append(f"[{p['name']}] {measurements[i]['date'].strftime('%Y-%m-%d')} 키 보정: 빈값 → {interp}cm (보간)")
                measurements[i]['height'] = interp

    # ─── Interpolate missing weights ───
    for i in range(len(measurements)):
        if measurements[i]['weight'] is None:
            prev_w, prev_d = None, None
            next_w, next_d = None, None
            for j in range(i-1, -1, -1):
                if measurements[j]['weight'] is not None:
                    prev_w = measurements[j]['weight']
                    prev_d = measurements[j]['date']
                    break
            for j in range(i+1, len(measurements)):
                if measurements[j]['weight'] is not None:
                    next_w = measurements[j]['weight']
                    next_d = measurements[j]['date']
                    break
            if prev_w and next_w and prev_d and next_d:
                total_days = (next_d - prev_d).days
                curr_days = (measurements[i]['date'] - prev_d).days
                ratio = curr_days / total_days if total_days > 0 else 0.5
                interp = round(prev_w + (next_w - prev_w) * ratio, 1)
                changes_log.append(f"[{p['name']}] {measurements[i]['date'].strftime('%Y-%m-%d')} 몸무게 보정: 빈값 → {interp}kg (보간)")
                measurements[i]['weight'] = interp

    # ─── Fix height anomalies (must always increase) ───
    for i in range(1, len(measurements)):
        if measurements[i]['height'] is not None and measurements[i-1]['height'] is not None:
            if measurements[i]['height'] < measurements[i-1]['height']:
                old_val = measurements[i]['height']
                # Use prev and next valid increasing value
                prev_h = measurements[i-1]['height']
                next_h = None
                for j in range(i+1, len(measurements)):
                    if measurements[j]['height'] is not None and measurements[j]['height'] > prev_h:
                        next_h = measurements[j]['height']
                        next_d = measurements[j]['date']
                        break
                if next_h:
                    total_days = (next_d - measurements[i-1]['date']).days
                    curr_days = (measurements[i]['date'] - measurements[i-1]['date']).days
                    ratio = curr_days / total_days if total_days > 0 else 0.5
                    new_val = round(prev_h + (next_h - prev_h) * ratio, 1)
                else:
                    new_val = round(prev_h + 0.5, 1)  # minimal increase
                changes_log.append(f"[{p['name']}] {measurements[i]['date'].strftime('%Y-%m-%d')} 키 보정: {old_val}→{new_val}cm (키 감소 수정)")
                measurements[i]['height'] = new_val

    # ─── Fix weight anomalies where clearly wrong (optional, less strict) ───
    # Weight can fluctuate, but if (?) was noted and it decreased, fix it
    # We'll leave weight as-is unless it's clearly wrong

    # ─── Calculate AI predicted height ───
    for m in measurements:
        if m['ba_months'] is not None and m['height'] is not None:
            m['ai_predicted'] = predict_adult_height(p['gender'], m['ba_months'], m['height'])
        else:
            m['ai_predicted'] = None

    p['measurements'] = measurements
    results.append(p)

# ─── Print changes ───
print("=== 변경 사항 ===")
for c in changes_log:
    print(c)

# ─── Create Excel ───
wb = Workbook()
ws = wb.active
ws.title = "정리된 케이스"

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='667eea')
patient_fill = PatternFill('solid', fgColor='E8ECFF')
patient_font = Font(name='Arial', bold=True, size=11)
data_font = Font(name='Arial', size=10)
corrected_font = Font(name='Arial', size=10, color='0000FF')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Headers (원본 컬럼 구조 보존 + AI예상키 추가)
headers = ['카테고리', '부모 ID', '차트번호', '성함', '생년월일', '성별', '아빠 키', '엄마 키', '목표키 (THT)', '상담 메모',
           '날짜', '키(cm)', '몸무게(kg)', '실제나이', '뼈나이', '뼈예상키 (THT/PAH)', 'AI예상키(계산)',
           '치료 메모1', '치료 메모2', '치료 후기', '유투브 링크']

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Column widths
widths = [20, 8, 10, 10, 12, 5, 7, 7, 8, 15, 12, 8, 10, 10, 10, 14, 14, 25, 25, 40, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data
row_num = 2
for p in results:
    # Patient header row
    first_m = p['measurements'][0] if p['measurements'] else None

    for mi, m in enumerate(p['measurements']):
        r = row_num

        if mi == 0:
            # Col 1~10: patient info (only on first measurement row)
            for ci, val in [(1, p['category']), (2, p['parent_id']), (3, p['chart_no']),
                            (4, p['name']),
                            (5, p['dob'].strftime('%Y-%m-%d') if p['dob'] else ''),
                            (6, p['gender']), (7, p['dad_h']), (8, p['mom_h']),
                            (9, p['target']), (10, p['consult_memo'])]:
                ws.cell(r, ci, val).font = patient_font
                ws.cell(r, ci).fill = patient_fill

        # Col 11~21: measurement data
        ws.cell(r, 11, m['date'].strftime('%Y-%m-%d')).font = data_font
        ws.cell(r, 12, m['height']).font = data_font
        ws.cell(r, 13, m['weight']).font = data_font
        ws.cell(r, 14, m['age_str']).font = data_font
        ws.cell(r, 15, m['ba_str']).font = data_font
        ws.cell(r, 16, m['original_pah']).font = data_font

        if m['ai_predicted']:
            ws.cell(r, 17, m['ai_predicted']).font = Font(name='Arial', size=10, bold=True, color='667eea')

        ws.cell(r, 18, m['memo1']).font = data_font
        ws.cell(r, 19, m['memo2']).font = data_font

        if mi == 0:
            ws.cell(r, 20, p['review']).font = data_font
            ws.cell(r, 20).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(r, 21, p['youtube']).font = data_font

        # Apply borders
        total_cols = len(headers)
        for col in range(1, total_cols + 1):
            ws.cell(r, col).border = thin_border
            ws.cell(r, col).alignment = Alignment(vertical='top', wrap_text=True) if col >= 18 else Alignment(vertical='center')

        row_num += 1

    # Empty separator row
    row_num += 1

# ─── Changes log sheet ───
ws2 = wb.create_sheet("변경 로그")
ws2.cell(1, 1, "변경 사항").font = Font(name='Arial', bold=True, size=12)
ws2.column_dimensions['A'].width = 80
for i, c in enumerate(changes_log, 2):
    ws2.cell(i, 1, c).font = Font(name='Arial', size=10)

# ─── Summary sheet ───
ws3 = wb.create_sheet("환자 요약")
summary_headers = ['이름', '성별', '생년월일', '아빠키', '엄마키', '초진일', '최근일',
                   '초진키', '최근키', '키성장', '초진뼈예상키', '최근AI예상키', '치료기간']
for col, h in enumerate(summary_headers, 1):
    cell = ws3.cell(1, col, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for i, p in enumerate(results, 2):
    ms = p['measurements']
    if not ms:
        continue
    first, last = ms[0], ms[-1]
    growth = round(last['height'] - first['height'], 1) if first['height'] and last['height'] else ''
    duration_days = (last['date'] - first['date']).days
    duration_str = f"{duration_days // 365}년 {(duration_days % 365) // 30}개월"

    first_pah = first['original_pah'] if first['original_pah'] else (first['ai_predicted'] or '')
    last_ai = last['ai_predicted'] or ''

    vals = [p['name'], p['gender'], p['dob'].strftime('%Y-%m-%d') if p['dob'] else '',
            p['dad_h'], p['mom_h'],
            first['date'].strftime('%Y-%m-%d'), last['date'].strftime('%Y-%m-%d'),
            first['height'], last['height'], growth,
            first_pah, last_ai, duration_str]
    for col, v in enumerate(vals, 1):
        ws3.cell(i, col, v).font = data_font

for col in range(1, 14):
    ws3.column_dimensions[get_column_letter(col)].width = 14

# Freeze panes
ws.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'

outpath = 'cases/치료 후기 케이스 정리 (정리완료).xlsx'
wb.save(outpath)
print(f"\n=== 저장 완료: {outpath} ===")
print(f"총 {len(results)}명, 변경사항 {len(changes_log)}건")

# Print summary
print("\n=== 환자별 AI 예상키 ===")
for p in results:
    print(f"\n● {p['name']} ({p['gender']}, DOB: {p['dob'].strftime('%Y-%m-%d') if p['dob'] else '?'})")
    for m in p['measurements']:
        ai = f"→ AI예상키: {m['ai_predicted']}cm" if m['ai_predicted'] else ""
        ba = f"(뼈나이: {m['ba_str']})" if m['ba_str'] else ""
        print(f"  {m['date'].strftime('%Y-%m-%d')} | 키:{m['height']}cm | {m['age_str']} | {ba} {ai}")
